# -*- coding: utf-8 -*-
# 用普通 python 运行（非 browser-use）。从 extract.py 输出的 vuex_raw.json 读取，
# 解码价格、解析销量/有效期/满减，按规格去重 + 按系列归并，产出
#   <BRAND>_热销统计_系列合并_<日期>.xlsx  （三表：热销系列 / TOP10最低价3档位 / 产品明细）
#   <BRAND>_热销采购分析_<日期>.html
#   sales_history.json  （当日 TOP10 快照，供趋势分析）
#   ysb_images/         （商品图）
#
# 用法（无参数时沿用默认，向后兼容）：
#   python process.py
#   python process.py --brand 妇炎洁 --input vuex_raw.json --detail detail_data.json
import json, base64, re, os, datetime, urllib.request, sys, io, argparse, time
import openpyxl
import ysb_parser

_RUN_START = time.time()  # 记录 process.py 启动时间，用于计算运行耗时
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ====================== 默认配置（无 CLI 参数时使用）======================
DEFAULT_BRAND = "妇炎洁"
DEFAULT_INPUT_JSON = "vuex_raw.json"
DEFAULT_DETAIL = "detail_data.json"
DEFAULT_IMG_DIR = "ysb_images"
# ========================================================================

_args = argparse.ArgumentParser(description="药帮采集结果解析与出表")
_args.add_argument("--brand", default=DEFAULT_BRAND, help="品牌名（默认 %s）" % DEFAULT_BRAND)
_args.add_argument("--input", default=DEFAULT_INPUT_JSON, help="vuex_raw.json 路径（默认 %s）" % DEFAULT_INPUT_JSON)
_args.add_argument("--detail", default=DEFAULT_DETAIL, help="detail_data.json 路径（默认 %s）" % DEFAULT_DETAIL)
_args.add_argument("--img-dir", default=DEFAULT_IMG_DIR, help="图片目录（默认 %s）" % DEFAULT_IMG_DIR)
_args.add_argument("--output-xlsx", default=None, help="xlsx 路径，为空则用 <BRAND>_热销统计_系列合并_<日期>.xlsx")
_args.add_argument("--output-html", default=None, help="html 路径，为空则用 <BRAND>_热销采购分析_<日期>.html")
args = _args.parse_args()

BRAND = args.brand
INPUT_JSON = args.input
DETAIL_JSON = args.detail
IMG_DIR = args.img_dir
OUTPUT_XLSX = args.output_xlsx
OUTPUT_HTML = args.output_html
# ── 配置记忆：记录上次运行参数到 profile.json，下次运行时自动建议 ──
_SKILL_DIR_MEM = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PROFILE_MEM = os.path.join(_SKILL_DIR_MEM, "profile.json")
try:
    with open(_PROFILE_MEM, encoding="utf-8") as _f:
        _mem = json.load(_f)
except Exception:
    _mem = {}
_last_cfg = _mem.get("last_config", {})
_last_brand = _last_cfg.get("brand", "")
if _last_brand and _last_brand != BRAND:
    print("[grow] 上次运行品牌: %s（当前: %s）" % (_last_brand, BRAND))
if _last_cfg:
    _lp = _last_cfg.get("pages", "?")
    print("[grow] 上次配置: brand=%s, pages=%s" % (_last_brand or "—", _lp))
# 通用品牌/前缀剥离正则（用于规格去重与系列归并，去掉品牌名、包邮、®™ 等）
BRAND_RE = re.compile(r'包邮|云闪付专享|®|™|' + re.escape(BRAND))
# 商品详情页链接、商品类型判断 — 统一引用 ysb_parser 公共模块
PRODUCT_URL_GROUP = ysb_parser.PRODUCT_URL_GROUP
PRODUCT_URL_REGULAR = ysb_parser.PRODUCT_URL_REGULAR
is_group_buy_name = ysb_parser.is_group_buy_name
# =================================================

raw = json.load(open(INPUT_JSON, encoding="utf-8-sig"))

# ====================== 串号去重 ======================
# 多个 Vuex 条目匹配同一张卡片时 domText 完全相同但 provider_name 不同，
# 导致同一销量被重复计入（虚高 2–20×）。按 domText 去重，每张唯一卡片只保留一条。
_seen_dom = set()
_deduped = []
for _it in raw:
    _dom = (_it.get("domText") or "").strip()
    if _dom and len(_dom) > 50:
        if _dom in _seen_dom:
            continue
        _seen_dom.add(_dom)
    _deduped.append(_it)
if len(_deduped) < len(raw):
    print("串号去重: %d → %d（去除 %d 条重复卡片匹配）" % (len(raw), len(_deduped), len(raw) - len(_deduped)))
raw = _deduped


# 价格解码、销量解析 — 统一引用 ysb_parser 公共模块
decode_price = ysb_parser.decode_price
parse_sales = ysb_parser.parse_sales


def window(dom, name):
    if not dom:
        return ""
    p = dom.find(name)
    if p == -1:
        p = dom.find(name[:18]) if name else -1
    if p == -1:
        return dom[:500]
    return dom[max(0, p - 250): p + len(name) + 420]


def parse_expiry(dom):
    if not dom:
        return ""
    m = re.search(r'有效期\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2})', dom)
    return m.group(1) if m else ""


def parse_tiers(dom, unit):
    if not dom:
        return []
    out = []
    for m in re.finditer(r'满\s*(\d+)\s*[盒瓶件袋包粒片套罐]?\s*享\s*([\d.]+)\s*元/[盒瓶件袋包粒片套罐]', dom):
        out.append((int(m.group(1)), m.group(2)))
    return out


def _normalize_product(name, spec):
    """提取产品核心属性（类型+风味+重量），忽略标题中的包邮/起购/品牌/重复词。
    不依赖 BRAND_RE 剥离品牌——直接提取属性，避免品牌名含产品类型词（如'云南白药牙膏'）时误删'牙膏'。"""
    s = (name or "") + " " + (spec or "")
    # 去前缀：N盒包邮 / N盒起购 / N支包邮
    s = re.sub(r'\d+[盒瓶件袋包粒片套罐支]+\s*(包邮|起购)', ' ', s)
    # 去【】
    s = re.sub(r'【[^】]*】', ' ', s)
    # 去括号但保留内容
    s = re.sub(r'[（(]', ' ', s)
    s = re.sub(r'[）)]', ' ', s)
    # 提取重量（如 230g, 215g, 145g+30g）— 去重避免重复词导致不同键
    weights = list(dict.fromkeys(re.findall(r'\d+\s*g', s)))
    weight = '+'.join(w.replace(' ', '') for w in weights) if weights else ''
    # 提取风味（按长度降序匹配，避免短词误命中）
    flavor = ''
    for f in ['薄荷清爽香型', '薄荷清爽型', '留兰香型', '高地留兰香型', '益优冰柠', '冰柠', '留兰']:
        if f in s:
            flavor = f
            break
    # 提取产品类型
    ptype = ''
    if '金口健' in s:
        ptype = '金口健牙膏'
    elif '牙膏' in s:
        ptype = '牙膏'
    key = ptype + flavor
    return (key + "|" + weight) if (key or weight) else (name or "")


def norm_key(name, spec):
    """Sheet3 去重键：用同一归一化逻辑。"""
    return _normalize_product(name, spec)


def product_key(name, spec):
    """产品分组键：按「类型+风味|重量」分组，不同规格=不同产品。"""
    return _normalize_product(name, spec)


# ---------- 解析每条 listing ----------
listings = []
for it in raw:
    price = decode_price(it.get("priceToken"))
    if price is None:
        continue
    dom = window(it.get("domText", ""), it.get("drugname", ""))
    rec = {
        "name": (it.get("drugname") or "").strip(),
        "spec": (it.get("specification") or "").strip(),
        "unit": it.get("unit") or "",
        "min_order": it.get("minamount"),
        "provider": it.get("provider_name") or "",
        "price": float(price),
        "sales": parse_sales(dom),
        "expiry": parse_expiry(dom),
        "tiers": parse_tiers(dom, it.get("unit") or ""),
        "img": it.get("drugimageurl") or "",
        "wholesaleid": str(it.get("wholesaleid") or "").strip(),
        "key": norm_key(it.get("drugname"), it.get("specification")),
        "product": product_key(it.get("drugname"), it.get("specification")),
    }
    listings.append(rec)
print("listings:", len(listings))

# ---------- 按产品(名+规格)分组（不同规格=不同产品）----------
product_groups = {}
for r in listings:
    product_groups.setdefault(r["product"], []).append(r)


def product_summary(recs):
    # 防虚高：同一独立商品(名+供应商+规格)可能被同商家挂成多个报价档位(listing)，
    # 每个 listing 有各自真实「已拼」销量。产品总销量 = 每款独立商品取「最高已拼」后求和。
    prod_best = {}
    for x in recs:
        pk = (x["name"], x["provider"], x["spec"])
        if pk not in prod_best or x["sales"] > prod_best[pk]:
            prod_best[pk] = x["sales"]
    total = sum(prod_best.values())
    rep = max(recs, key=lambda x: x["sales"])
    suppliers = set(x["provider"] for x in recs if x["provider"])
    expiries = [x["expiry"] for x in recs if x["expiry"]]

    # ---- 聚合所有阶梯价（去重：同一起订量+价格+单位只保留一个）----
    tiers = []
    seen_tier = set()
    for x in recs:
        tk = (x["min_order"], round(x["price"], 2), x["unit"])
        if tk not in seen_tier and x["min_order"] is not None:
            seen_tier.add(tk)
            tiers.append({"min_order": x["min_order"], "price": round(x["price"], 2), "unit": x["unit"]})
    # 按起订量升序排列（少量→大量）
    tiers.sort(key=lambda t: (t["min_order"] or 0))

    # ---- 从详情页数据中提取近7天销量和大单统计（取代表报价的详情值）----
    rep_wid = rep.get("wholesaleid", "")
    detail = {}
    if rep_wid and rep_wid in detail_map:
        detail = detail_map[rep_wid]

    return {
        "product": recs[0]["product"],
        "rep_name": rep["name"],
        "rep_spec": rep["spec"],
        "rep_img": rep["img"],
        "rep_wid": rep_wid,
        # 代表商品(销量最高那家)自身的真实单价/起订量 —— 与下方商品ID链接严格对应
        "rep_price": rep["price"],
        "rep_min_order": rep["min_order"],
        # 全网最低价：同一产品家族下所有商户报价的最小值（供采购比价参考，单独标注，不等同于链接商品价）
        "lowest_price": min(x["price"] for x in recs),
        "min_order_min": min((x["min_order"] for x in recs if x["min_order"]), default=None),
        "total_sales": total,
        "max_sales": max(x["sales"] for x in recs),
        "n_offers": len(recs),
        "n_products": len(prod_best),
        "n_suppliers": len(suppliers),
        "expiries": expiries,
        # 完整阶梯价列表（按起订量升序）
        "tiered_prices": tiers,
        # 详情页新增指标
        "last_7_days_sales": detail.get("last_7_days_sales"),
        "large_orders_count": detail.get("large_orders_count"),
        "large_orders_total_qty": detail.get("large_orders_total_qty"),
    }


# ---------- 载入详情页权威数据（已付款件数/参团店数/采购笔数/资质/最近采购/近7天/大单）----------
try:
    detail_map = json.load(open(DETAIL_JSON, encoding="utf-8"))
except Exception:
    detail_map = {}


product_list = [product_summary(rs) for rs in product_groups.values()]
product_list.sort(key=lambda s: s["total_sales"], reverse=True)
top_products = product_list[:10]
print("products count:", len(product_list), "| top1:", top_products[0]["product"], top_products[0]["total_sales"])

# ---------- 数据质量校验 ----------
for s in product_list:
    recs = product_groups.get(s["product"], [])
    if len(recs) > 1:
        sales_vals = [x["sales"] for x in recs if x["sales"] > 0]
        if sales_vals:
            from collections import Counter
            sc = Counter(sales_vals)
            dups = {k: v for k, v in sc.items() if v > 1 and k > 1000}
            if dups:
                print("⚠ 数据质量警告: 产品「%s」内仍有重复销量 %s（可能残留串号）" % (s["rep_name"][:30], dups))


# 取某系列的 3 个最低"价格档位"（不同价格水平，每档取该价位销量最高的一家代表）
def three_tiers(recs):
    offers = []
    for x in recs:
        offers.append({"price": x["price"], "qty": x["min_order"], "label": "起订%s%s" % (x["min_order"], x["unit"]),
                       "expiry": x["expiry"], "sales": x["sales"], "provider": x["provider"], "wid": x["wholesaleid"]})
        for (q, p) in x["tiers"]:
            offers.append({"price": float(p), "qty": q, "label": "满%d%s" % (q, x["unit"]),
                           "expiry": x["expiry"], "sales": x["sales"], "provider": x["provider"], "wid": x["wholesaleid"]})
    offers.sort(key=lambda o: (o["price"], -o["sales"]))
    levels = {}
    for o in offers:
        p = round(o["price"], 2)
        if p not in levels or o["sales"] > levels[p]["sales"]:
            levels[p] = o
    return sorted(levels.values(), key=lambda o: o["price"])[:3]


# ---------- 图片下载 ----------
os.makedirs(IMG_DIR, exist_ok=True)
ua = {"User-Agent": "Mozilla/5.0"}


def dl(img_url, path):
    if not img_url:
        return ""
    if os.path.exists(path) and os.path.getsize(path) > 500:
        return path
    try:
        req = urllib.request.Request(img_url, headers=ua)
        data = urllib.request.urlopen(req, timeout=25).read()
        # 内存中压缩为缩略图（不落临时文件，避免触发批量删除安全策略）
        try:
            im = PILImage.open(io.BytesIO(data)).convert("RGB")
            im.thumbnail((120, 120))
            im.save(path, "JPEG", quality=82)
        except Exception:
            open(path, "wb").write(data)
        return path
    except Exception:
        return ""


# ---------- Excel ----------
wb = openpyxl.Workbook()
HEAD_FILL = PatternFill("solid", fgColor="4472C4")
def style_header(ws, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def link_cell(cell, wid, name=""):
    """把单元格设为 wholesaleid 文本，并附商品详情页超链接。
    根据商品名自动选择拼团/普通路由参数。"""
    if wid:
        cell.value = wid
        url = PRODUCT_URL_GROUP if is_group_buy_name(name) else PRODUCT_URL_REGULAR
        cell.hyperlink = url % wid
        cell.font = Font(color="0563C1", underline="single")
    else:
        cell.value = ""


# Sheet1: 热销产品（按规格分列，不同规格=不同产品）
ws = wb.active
ws.title = "热销产品"
hdr1 = ["排名", "图片", "产品名称", "规格", "本商品单价(元)", "起订量", "总销量", "供应商数", "全网最低价",
        "有效期示例", "商品ID"]
ws.append(hdr1)
style_header(ws, len(hdr1))
for i, s in enumerate(product_list, 1):
    r = i + 1
    ws.cell(row=r, column=1, value=i)
    ip = dl(s["rep_img"], "%s/prod_%d.jpg" % (IMG_DIR, i))
    if ip:
        try:
            im = XLImage(ip); im.width = 46; im.height = 46; ws.add_image(im, "B%d" % r)
        except Exception:
            pass
    ws.cell(row=r, column=3, value=s["rep_name"]).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(row=r, column=4, value=s["rep_spec"]).alignment = Alignment(wrap_text=True, vertical="center")
    # 本商品单价 = 代表商品(链接那家)自身真实价，与「商品ID」链接严格对应
    ws.cell(row=r, column=5, value=s["rep_price"])
    ws.cell(row=r, column=6, value=s["rep_min_order"])
    ws.cell(row=r, column=7, value=s["total_sales"])
    ws.cell(row=r, column=8, value=s["n_suppliers"])
    # 全网最低价：同产品家族下所有商户报价最小值（比价参考，不等于链接商品价）
    ws.cell(row=r, column=9, value=s["lowest_price"])
    ws.cell(row=r, column=10, value=(s["expiries"][0] if s["expiries"] else "")).alignment = Alignment(wrap_text=True, vertical="center")
    link_cell(ws.cell(row=r, column=11), s["rep_wid"], s["rep_name"])
    ws.row_dimensions[r].height = 50
for ci, w in enumerate([6, 10, 30, 22, 13, 9, 12, 9, 12, 16, 14], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = "A2"

# Sheet2: TOP10 最低价档位（产品内 3 个最低报价）
ws2 = wb.create_sheet("TOP10最低价档位")
hdr2 = ["产品排名", "产品名称", "规格", "档位", "价格(元)", "起订/满量", "有效期", "该价销量", "供应商", "商品ID"]
ws2.append(hdr2)
style_header(ws2, len(hdr2))
rr = 2
for ri, s in enumerate(top_products, 1):
    recs = product_groups[s["product"]]
    three = three_tiers(recs)
    for ti, o in enumerate(three):
        ws2.cell(row=rr, column=1, value=ri)
        ws2.cell(row=rr, column=2, value=s["rep_name"]).alignment = Alignment(wrap_text=True, vertical="center")
        ws2.cell(row=rr, column=3, value=s["rep_spec"]).alignment = Alignment(wrap_text=True, vertical="center")
        ws2.cell(row=rr, column=4, value=["最低", "次低", "次次低"][ti])
        ws2.cell(row=rr, column=5, value=o["price"])
        ws2.cell(row=rr, column=6, value=o["label"])
        ws2.cell(row=rr, column=7, value=o["expiry"])
        ws2.cell(row=rr, column=8, value=o["sales"])
        ws2.cell(row=rr, column=9, value=o["provider"]).alignment = Alignment(wrap_text=True, vertical="center")
        link_cell(ws2.cell(row=rr, column=10), o["wid"], s["rep_name"])
        rr += 1
for ci, w in enumerate([8, 26, 14, 8, 10, 16, 14, 10, 22, 14], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = "A2"

# Sheet3: 产品明细（去重，按规格，附产品列）
ws3 = wb.create_sheet("产品明细(去重)")
prod_groups = {}
for r in listings:
    prod_groups.setdefault(r["key"], []).append(r)
prod_rank = sorted(prod_groups.items(), key=lambda kv: max(x["sales"] for x in kv[1]), reverse=True)
deduped = []
for key, rs in prod_rank:
    best = max(rs, key=lambda x: x["sales"])
    d = dict(best); d["_nlist"] = len(rs); deduped.append(d)
hdr3 = ["排名", "图片", "标题", "规格", "单价(元)", "单位", "起订量", "销量", "供应商", "有效期", "商品ID"]
ws3.append(hdr3)
style_header(ws3, len(hdr3))
for i, d in enumerate(deduped, 1):
    r = i + 1
    ws3.cell(row=r, column=1, value=i)
    ip = d["img"] and dl(d["img"], "%s/det_%d.jpg" % (IMG_DIR, i))
    if ip:
        try:
            im = XLImage(ip); im.width = 46; im.height = 46; ws3.add_image(im, "B%d" % r)
        except Exception:
            pass
    ws3.cell(row=r, column=3, value=d["name"]).alignment = Alignment(wrap_text=True, vertical="center")
    ws3.cell(row=r, column=4, value=d["spec"]).alignment = Alignment(wrap_text=True, vertical="center")
    ws3.cell(row=r, column=5, value=d["price"])
    ws3.cell(row=r, column=6, value=d["unit"])
    ws3.cell(row=r, column=7, value=d["min_order"])
    ws3.cell(row=r, column=8, value=d["sales"])
    ws3.cell(row=r, column=9, value=d["provider"]).alignment = Alignment(wrap_text=True, vertical="center")
    ws3.cell(row=r, column=10, value=d["expiry"])
    link_cell(ws3.cell(row=r, column=11), d["wholesaleid"], d["name"])
    ws3.row_dimensions[r].height = 50
for ci, w in enumerate([6, 10, 40, 26, 10, 6, 8, 10, 22, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = "A2"

today = datetime.date.today().strftime("%Y-%m-%d")
fn = OUTPUT_XLSX or ("%s_热销统计_系列合并_%s.xlsx" % (BRAND, today))
wb.save(fn)
print("saved", fn)

# ---------- 历史快照（产品级 TOP10）----------
hist_path = "sales_history.json"
try:
    hist = json.load(open(hist_path, encoding="utf-8"))
except Exception:
    hist = []
snap = {"date": today, "top10": []}
for ri, s in enumerate(top_products, 1):
    recs = product_groups[s["product"]]
    pts = []
    for x in recs:
        pts.append(x["price"])
        for (q, p) in x["tiers"]:
            pts.append(float(p))
    pts.sort()
    snap["top10"].append({
        "rank": ri, "name": s["rep_name"], "spec": s["rep_spec"],
        "total_sales": s["total_sales"], "n_offers": s["n_offers"],
        "lowest3_prices": [round(pts[i], 2) for i in range(min(3, len(pts)))],
    })
hist = [h for h in hist if h["date"] != today]
hist.append(snap)
json.dump(hist, open(hist_path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("appended snapshot -> history days:", len(hist))

# ---------- 验证输出 ----------
print("\n=== TOP10 产品（按总销量，阶梯价 + 3 个不同价格档位）===")
for ri, s in enumerate(top_products, 1):
    three = three_tiers(product_groups[s["product"]])
    t3 = [o["price"] for o in three]
    tp = s.get("tiered_prices", [])
    tier_str = ", ".join("%s%s/¥%.2f" % (t["min_order"], t["unit"], t["price"]) for t in tp[:5])
    if len(tp) > 5:
        tier_str += " ... (共%d档)" % len(tp)
    l7d = s.get("last_7_days_sales")
    lc = s.get("large_orders_count")
    print("#%d %s %s | 总销量=%d 报价数=%d | 阶梯价=[%s] | 3档位=%s | 近7天=%s 大单=%d笔"
        % (ri, s["rep_name"], s["rep_spec"], s["total_sales"], s["n_offers"], tier_str,
           ",".join("%.2f" % p for p in t3), l7d or "—", lc or 0))

# ---------- 生成 HTML（可钻取：首页排名 + 档位页 10 链接，整合详情页权威销量）----------
json.dump(product_list, open("series.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)


def build_offers(recs):
    """每个产品取最多 10 个去重报价（按价升序），并附详情页权威销量字段。"""
    offers, seen = [], set()
    for x in recs:
        wid = x.get("wholesaleid")
        if wid in seen:
            continue
        seen.add(wid)
        det = detail_map.get(str(wid), {})
        offers.append({
            "wid": wid,
            "name": x.get("name", ""),
            "price": x["price"],
            "minOrder": x["min_order"], "unit": x["unit"],
            "provider": x["provider"], "expiry": (det.get("expiry_date") or x.get("expiry") or ""),
            "sales": x["sales"],
            "paidUnits": det.get("paid_units"), "paidUnit": det.get("paid_unit"),
            "stores": det.get("stores_joined"), "records": det.get("purchase_records"),
            "manufacturer": det.get("manufacturer"), "approvalNo": det.get("approval_no"),
            "produceDate": det.get("produce_date"),
            "recent": det.get("recent_purchases", []),
            "detailMinQty": det.get("min_qty"), "detailMinUnit": det.get("min_unit"),
            # 每个链接自己的近7天 / 大单数据（来自该 wholesaleid 的详情页）
            "last7d": det.get("last_7_days_sales"),
            "largeCount": det.get("large_orders_count"),
            "largeQty": det.get("large_orders_total_qty"),
            # 每个链接自己的阶梯价（该 wholesaleid 自身的 起订量→价格，不共用产品级列表）
            "tiers": [{"min_order": x["min_order"], "price": round(x["price"], 2), "unit": x["unit"]}]
                     if x["min_order"] is not None else [],
        })
    offers.sort(key=lambda o: (o["price"], -(o["paidUnits"] or 0)))
    return offers[:10]


def assign_tier(offers):
    """价格分位三档 + 销量标签（替代 最低/次低/次次低）。"""
    if not offers:
        return
    # 报价不足 3 个时无价格分位意义，统一归主流走量档
    if len(offers) < 3:
        for o in offers:
            o["tier"] = "主流走量档"
    else:
        prices = sorted(o["price"] for o in offers)
        n = len(prices)
        q1 = prices[max(0, n // 4 - 1)]
        q3 = prices[min(n - 1, 3 * n // 4)]
        for o in offers:
            p = o["price"]
            o["tier"] = "源头直供档" if p <= q1 else ("精选优价档" if p > q3 else "主流走量档")
    soff = sorted(offers, key=lambda o: (o["paidUnits"] if o["paidUnits"] is not None else -1), reverse=True)
    for idx, o in enumerate(soff):
        o["stag"] = "—" if o["paidUnits"] is None else ("爆款领跑" if idx == 0 else ("稳健供货" if idx < len(soff) // 2 else "长尾备选"))


products_out = []
for s in product_list:
    recs = product_groups.get(s["product"], [])
    offers = build_offers(recs)
    assign_tier(offers)
    # 代表报价：从该产品【全部报价】里挑详情页已付款最高者（不在仅展示的 10 个内也行）
    rep, best = None, -1
    for x in recs:
        det = detail_map.get(str(x.get("wholesaleid")), {})
        pu = det.get("paid_units")
        m = pu if pu is not None else (x["sales"] or 0)
        if m > best:
            best, rep = m, {"paidUnits": pu}
    sales_metric = (rep["paidUnits"] if rep and rep["paidUnits"] is not None else (s["total_sales"] or s["max_sales"] or 0))
    sales_src = "详情页" if (rep and rep["paidUnits"] is not None) else "列表"
    prices = [o["price"] for o in offers if o["price"]]
    avg = round(sum(prices) / len(prices), 2) if prices else s["lowest_price"]
    products_out.append({
        "id": 0, "name": s["rep_name"], "spec": s["rep_spec"], "repImg": s["rep_img"], "repWid": s["rep_wid"],
        "avg": avg, "nSup": s["n_suppliers"], "minP": s["lowest_price"],
        "expiry": (s["expiries"][0] if s["expiries"] else ""),
        "sales": sales_metric, "salesSrc": sales_src, "offers": offers,
        # 阶梯价（所有档位）
        "tieredPrices": s.get("tiered_prices", []),
        # 近7天 / 大单
        "last7d": s.get("last_7_days_sales"),
        "largeCount": s.get("large_orders_count"),
        "largeQty": s.get("large_orders_total_qty"),
    })
# 排名：按权威销量降序（列表兜底）
products_out.sort(key=lambda p: (p["sales"] or 0), reverse=True)
# HTML 图片直接用原始 URL，无需下载（浏览器在线加载，onerror 自动隐藏失效图）
# Excel 图片仍需下载（openpyxl 要求本地文件路径）
for i, p in enumerate(products_out, 1):
    p["id"] = i
    p["rank"] = i
    p["img"] = p.get("repImg") or ""

HTML = r"""<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>__BRAND__ 热销采购分析</title>
<style>
* { box-sizing:border-box; }
body { font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif; margin:0; background:#f5f7fa; color:#222; }
.wrap { max-width:1240px; margin:0 auto; padding:22px 18px 60px; }
h1 { font-size:22px; margin:0 0 4px; }
.sub { color:#888; font-size:13px; margin:6px 0 12px; }
.note { background:#eef7ff; border:1px solid #bae0ff; color:#07508c; padding:10px 14px; border-radius:8px; font-size:12.5px; margin-bottom:18px; line-height:1.75; }
.note b { color:#06407a; }
table { width:100%; border-collapse:collapse; background:#fff; border-radius:10px; overflow:hidden; box-shadow:0 1px 4px rgba(0,0,0,.06); }
th,td { padding:9px 11px; text-align:left; border-bottom:1px solid #eef0f3; vertical-align:middle; font-size:13px; }
th { background:#f7f9fc; color:#555; font-weight:600; white-space:nowrap; }
tr:hover { background:#fafcff; }
.rk { color:#bbb; width:34px; text-align:center; }
.img { width:70px; } .img img { width:58px; height:58px; object-fit:cover; border-radius:6px; border:1px solid #eee; }
.noimg { width:58px; height:58px; background:#f0f0f0; border-radius:6px; display:flex; align-items:center; justify-content:center; color:#aaa; font-size:11px; }
.name { font-weight:700; color:#1a1a1a; } .name a { color:#1a1a1a; text-decoration:none; } .name a:hover { color:#1677ff; text-decoration:underline; }
.spec { color:#888; font-size:12px; margin-top:2px; }
.price { color:#e4393c; font-weight:700; white-space:nowrap; } .mo { white-space:nowrap; color:#555; }
.sales { color:#c0392b; font-weight:700; white-space:nowrap; } .sup,.light { color:#555; }
.src { font-size:10px; color:#aaa; border:1px solid #ddd; border-radius:3px; padding:0 3px; margin-left:3px; }
.tier { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:700; color:#fff; }
.t1 { background:#16a34a; } .t2 { background:#1677ff; } .t3 { background:#ea8c2b; }
.stag { display:inline-block; margin-top:3px; font-size:10.5px; color:#666; }
.wid a { color:#0563C1; white-space:nowrap; }
.summary { display:flex; gap:12px; flex-wrap:wrap; margin:6px 0 16px; }
.summary > div { background:#fff; border:1px solid #eef0f3; border-radius:10px; padding:10px 16px; min-width:140px; box-shadow:0 1px 3px rgba(0,0,0,.04); }
.summary b { color:#1677ff; font-size:18px; display:block; margin-top:2px; }
.status-bar { background:#fff7e6; border:1px solid #ffd591; color:#ad6800; padding:9px 14px; border-radius:8px; font-size:12.5px; margin-bottom:14px; line-height:1.6; }
.status-bar b { color:#ad6800; font-size:14px; } .status-bar b.warn { color:#cf1322; }
tr[data-status="pending"] { background:#fafafa; color:#999; }
tr[data-status="pending"]:hover { background:#f5f5f5; }
.pending { color:#bbb; font-size:11px; }
.back { display:inline-block; margin:4px 0 10px; color:#1677ff; text-decoration:none; font-size:13px; }
.back:hover { text-decoration:underline; }
h2 { font-size:18px; margin:0 0 4px; } h3 { font-size:15px; margin:18px 0 8px; }
.recent { margin-top:14px; background:#fff; border:1px solid #eef0f3; border-radius:10px; padding:8px 14px; }
.recent summary { cursor:pointer; color:#1677ff; font-weight:600; font-size:13px; }
.rlist { list-style:none; margin:8px 0 2px; padding:0; max-height:260px; overflow:auto; }
.rlist li { display:flex; gap:14px; padding:4px 0; border-bottom:1px dashed #f0f0f0; font-size:12.5px; color:#444; }
.rlist li span:nth-child(1){ min-width:120px; } .rlist li span:nth-child(2){ min-width:90px; color:#999; }
.rlist li span:nth-child(3){ min-width:60px; font-weight:600; } .rlist li span:nth-child(4){ color:#1677ff; }
/* 首页表格 */
/* 详情页每行指标（阶梯价 / 近7天 / 大单） */
.tiers { font-size:11px; color:#555; max-width:200px; line-height:1.6; }
.tiers b { color:#e4393c; }
.metric { text-align:center; font-size:12px; }
.metric.large { color:#e4393c; }
</style></head><body><div class="wrap">
<h1>__BRAND__ 热销采购分析</h1>
<div class="sub">数据来源：药师帮 dian.ysbang.cn 批发搜索「__BRAND__」 · 抓取日期 __DATE__ · 共 __NP__ 个产品（不同规格为独立产品）</div>
<div id="view"></div>
</div>
<script>
const BRAND="__BRAND__";
const PRODUCTS = __DATA__;
const DETAIL_BASE="https://dian.ysbang.cn/#/drugInfo?wholesaleid=";
function detailUrl(w,name){return DETAIL_BASE+w+(name&&name.indexOf("包邮")>=0?"&isAssemble=true&scene=0":"&isAssemble=false&scene=1")+"&trafficType=1";}
function fmt(n){ if(n==null) return "—"; return n>=10000 ? (n/10000).toFixed(1)+"万" : Number(n).toLocaleString(); }
function tierCls(t){ return t==="源头直供档"?"t1":(t==="主流走量档"?"t2":"t3"); }
function renderIndex(){
  let rows = PRODUCTS.map(p=>{
    let st = fmt(p.sales) + (p.salesSrc==="列表"?'<span class="src">列表</span>':'');
    let img = p.img ? '<img src="'+p.img+'" onerror="this.style.display=\'none\'">' : '<div class="noimg">无图</div>';
    return '<tr><td class="rk">'+p.rank+'</td><td class="img">'+img+'</td>'
      +'<td class="name"><a href="#/p/'+p.id+'">'+p.name+'</a><div class="spec">'+p.spec+'</div></td>'
      +'<td class="price">￥'+p.minP+'</td><td>￥'+p.avg+'</td><td>'+p.nSup+' 家</td>'
      +'<td class="sales">'+st+'</td><td>'+(p.expiry||"—")+'</td></tr>';
  }).join("");
  let totalOffers = PRODUCTS.reduce((a,p)=>a+p.offers.length,0);
  let realOffers = PRODUCTS.reduce((a,p)=>a+p.offers.filter(o=>o.paidUnits!=null).length,0);
  let pending = totalOffers - realOffers;
  if(pending<0) pending=0;
  let bar = '<div class="status-bar">数据完整性：已采集 <b>'+realOffers+'</b> / '+totalOffers+' 条报价真实销量（详情页「已付款件数」） · 剩余 <b class="warn">'+pending+'</b> 条待补采（详情页抓取中途中断，跑稳后自动补全，当前以列表「已拼」量作参考）。</div>';
  document.getElementById("view").innerHTML = bar
    + '<div class="note"><b>用法：</b>首页按「销量排名 / 均价 / 供应商数 / 全网最低价」概览，<b>点击产品标题</b>进入该产品的档位页（10 个报价链接 + 详情页权威销量/近7天/大单统计）。<br>'
    +'<b>档位定义（按价格分位，已替代「最低/次低/次次低」）：</b>'
    +'<span class="tier t1">源头直供档</span> 价≤Q1（价格锚点）&nbsp; '
    +'<span class="tier t2">主流走量档</span> Q1&lt;价≤Q3（成交密集，含销量冠军）&nbsp; '
    +'<span class="tier t3">精选优价档</span> 价&gt;Q3（小批量/高服务/长效期）；并叠加销量标签 '
    +'<b>爆款领跑 / 稳健供货 / 长尾备选</b>。<br>'
    +'<b>权威销量 = 商品详情页「已付款件数」</b>（带单位，不同报价单位如盒/瓶不可相加，排名取代表报价单一值）；标「列表」者为搜索列表「已拼」量（仅供参考）。</div>'
    +'<table><thead><tr><th>排名</th><th>图</th><th>产品（点标题看档位）</th><th>全网最低价</th><th>均价</th><th>供应商数</th><th>销量(权威)</th><th>有效期</th></tr></thead><tbody>'+rows+'</tbody></table>';
}
function renderDetail(id){
  let p = PRODUCTS.find(x=>x.id==id);
  if(!p){ location.hash="#/"; return; }
  let off = p.offers;
  let paid = off.filter(o=>o.paidUnits!=null).map(o=>o.paidUnits);
  let stores = off.filter(o=>o.stores!=null).map(o=>o.stores);
  let recsN = off.filter(o=>o.records!=null).map(o=>o.records);
  let sum=a=>a.reduce((x,y)=>x+y,0);
  let maxP = paid.length?Math.max.apply(null,paid):null;
  let avgP = paid.length?Math.round(sum(paid)/paid.length):null;
  let maxUnit = maxP!=null ? off.find(o=>o.paidUnits===maxP).paidUnit : "";
  // 全产品阶梯价（汇总，仅放表头；每行用各链接自身阶梯价，不共用）
  let ladderHtml = (p.tieredPrices && p.tieredPrices.length)
    ? '<div class="tiers">'+p.tieredPrices.map(t=>t.min_order+t.unit+' <b>￥'+t.price+'</b>').join('<br>')+'</div>'
    : '—';
  let summary = '<div class="summary">'
    +'<div>报价数<b>'+off.length+'</b></div>'
    +'<div>权威销量(已付款)<b>'+(maxP!=null? (fmt(maxP)+'~'+fmt(avgP)+' '+maxUnit) : '—')+'</b></div>'
    +'<div>参团店铺合计<b>'+(sum(stores)? sum(stores).toLocaleString()+' 家':'—')+'</b></div>'
    +'<div>采购记录合计<b>'+(sum(recsN)? sum(recsN).toLocaleString()+' 笔':'—')+'</b></div>'
    +'<div>全产品阶梯价(汇总)<b style="font-size:12px;font-weight:500">'+ladderHtml+'</b></div></div>';
  let rows = off.map(o=>{
    let pu = o.paidUnits!=null ? (Number(o.paidUnits).toLocaleString()+' '+o.paidUnit)
      : '<span class="pending">待补采</span><br><span class="src">列表 '+fmt(o.sales)+'</span>';
    let mq = o.detailMinQty ? (o.detailMinQty+o.detailMinUnit+'起拼') : (o.minOrder+o.unit+'起');
    // 每行自己的近7天 / 大单
    let l7d = o.last7d!=null ? fmt(o.last7d) : '—';
    let lc = o.largeCount!=null ? o.largeCount : '—';
    let lq = o.largeQty!=null ? fmt(o.largeQty) : '—';
    // 该链接自身的阶梯价（不共用产品级列表）
    let ot = (o.tiers && o.tiers.length)
      ? '<div class="tiers">'+o.tiers.map(t=>t.min_order+t.unit+' <b>￥'+t.price+'</b>').join('<br>')+'</div>'
      : '—';
    return '<tr data-status="'+(o.paidUnits!=null?'real':'pending')+'"><td><span class="tier '+tierCls(o.tier)+'">'+o.tier+'</span><br><span class="stag">'+o.stag+'</span></td>'
      +'<td class="price">￥'+o.price+'</td><td class="mo">'+mq+'</td><td>'+(o.provider||"—")+'</td>'
      +'<td>'+(o.expiry||"—")+'</td><td class="sales">'+pu+'</td>'
      +'<td>'+(o.stores!=null? o.stores+' 家':"—")+'</td><td>'+(o.records!=null? o.records+' 笔':"—")+'</td>'
      +'<td class="wid"><a href="'+detailUrl(o.wid,o.name)+'" target="_blank">详情↗</a></td>'
      +'<td class="tiers">'+ot+'</td>'
      +'<td class="metric">'+l7d+'</td>'
      +'<td class="metric large">'+lc+'</td>'
      +'<td class="metric large">'+lq+'</td></tr>';
  }).join("");
  let rep = off.find(o=>o.recent && o.recent.length);
  let recent = "";
  if(rep){
    let items = rep.recent.map(r=>'<li><span>'+r.buyer+'</span><span>'+r.phone+'</span><span>'+r.qty+'</span><span>'+r.time+'</span></li>').join("");
    recent = '<details class="recent"><summary>最近采购明细（'+ (rep.provider||'') +' · 销量时序）</summary><ul class="rlist">'+items+'</ul></details>';
  }
  document.getElementById("view").innerHTML =
    '<a class="back" href="#/">← 返回全部产品</a>'
    +'<h2>'+p.name+' <span class="spec">'+p.spec+'</span></h2>'
    +'<div class="sub">全网最低价 ￥'+p.minP+' · 均价 ￥'+p.avg+' · '+p.nSup+' 家供应商 · 权威销量 '+fmt(p.sales)+'（'+p.salesSrc+'）</div>'
    +summary
    +'<h3>报价档位（'+off.length+' 个 · 按价升序，点「详情↗」进商品页）</h3>'
    +'<table><thead><tr><th>档位 / 标签</th><th>成团价</th><th>起拼</th><th>供应商</th><th>有效期</th><th>已付款(权威)</th><th>参团店</th><th>采购笔</th><th>链接</th><th>阶梯价(该链接)</th><th>近7天销量</th><th>大单≥50(笔)</th><th>大单总销量</th></tr></thead><tbody>'+rows+'</tbody></table>'
    +recent;
  window.scrollTo(0,0);
}
function route(){
  let h = location.hash;
  if(h.indexOf("#/p/")===0){ renderDetail(h.split("/")[2]); }
  else { renderIndex(); }
}
window.addEventListener("hashchange", route);
route();
</script></body></html>"""
html = (HTML.replace("__BRAND__", BRAND).replace("__DATE__", today)
        .replace("__NP__", str(len(products_out)))
        .replace("__DATA__", json.dumps(products_out, ensure_ascii=False)))
html_fn = OUTPUT_HTML or ("%s_热销采购分析_%s.html" % (BRAND, today))
open(html_fn, "w", encoding="utf-8").write(html)
print("generated", html_fn, "| products:", len(products_out))

# ---------- 成长记录（skill 自我积累经验 · 增强版）----------
# 记录内容：运行次数、采集量、品牌排行、数据质量、耗时、配置、趋势
# 优化点：同日去重、数据质量追踪、耗时记录、配置记忆、runs上限200
try:
    _SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _PROFILE = os.path.join(_SKILL_DIR, "profile.json")
    try:
        with open(_PROFILE, encoding="utf-8") as _f:
            _prof = json.load(_f)
    except Exception:
        _prof = {"version": 2, "first_run": today, "last_run": today,
                 "total_runs": 0, "total_items": 0, "brands": {}, "runs": []}

    # 升级旧版 profile
    if _prof.get("version", 1) < 2:
        _prof["version"] = 2
        _prof.setdefault("quality_history", [])
        _prof.setdefault("lessons", [])

    _now = datetime.datetime.now().isoformat(timespec="seconds")
    _elapsed = round(time.time() - _RUN_START, 1)
    _top1 = products_out[0] if products_out else {}

    # ── 数据质量统计 ──
    _total_wids = len(set(str(r.get("wholesaleid")) for r in listings if r.get("wholesaleid")))
    _detail_ok = 0
    _detail_pending = 0
    if detail_map:
        for _wid in (str(r.get("wholesaleid")) for r in listings if r.get("wholesaleid")):
            _dd = detail_map.get(_wid)
            if _dd and _dd.get("paid_units") is not None:
                _detail_ok += 1
            else:
                _detail_pending += 1
    _quality_rate = round(_detail_ok / _total_wids * 100, 1) if _total_wids > 0 else 0

    _run = {
        "date": _now,
        "brand": BRAND,
        "elapsed_sec": _elapsed,
        "result": {
            "items": len(listings),
            "products": len(product_list),
            "top1_name": _top1.get("name", ""),
            "top1_sales": _top1.get("sales", 0),
        },
        "quality": {
            "total_wids": _total_wids,
            "detail_collected": _detail_ok,
            "detail_pending": _detail_pending,
            "collection_rate": _quality_rate,
        },
        "config": {
            "brand": BRAND,
            "pages": None,  # extract.py 的参数，这里无法直接获取
            "input": os.path.basename(INPUT_JSON),
            "has_detail": bool(detail_map),
        },
    }

    # ── 同日同品牌去重：移除当天同品牌的旧记录，只保留最新一次 ──
    _prof["runs"] = [r for r in _prof["runs"]
                     if not (r.get("date", "").startswith(today) and r.get("brand") == BRAND)]
    _prof["runs"].append(_run)

    # ── runs 上限：保留最近 200 条 ──
    MAX_RUNS = 200
    if len(_prof["runs"]) > MAX_RUNS:
        _archived = _prof["runs"][:-MAX_RUNS]
        _prof.setdefault("archived_runs", []).extend(_archived)
        _prof["runs"] = _prof["runs"][-MAX_RUNS:]

    # ── 全局统计 ──
    _prof["total_runs"] += 1
    _prof["total_items"] += len(listings)
    _prof["last_run"] = today

    # ── 品牌级统计 ──
    _b = _prof["brands"].setdefault(BRAND, {
        "runs": 0, "first_run": today, "last_run": today,
        "last_result": {}, "total_items": 0,
        "best_quality": 0, "avg_elapsed": 0, "total_elapsed": 0,
    })
    _b["runs"] += 1
    _b["last_run"] = today
    _b["last_result"] = _run["result"]
    _b["total_items"] += len(listings)
    _b["total_elapsed"] += _elapsed
    _b["avg_elapsed"] = round(_b["total_elapsed"] / _b["runs"], 1)
    if _quality_rate > _b.get("best_quality", 0):
        _b["best_quality"] = _quality_rate

    # ── 数据质量历史（用于追踪完整性变化）──
    _prof.setdefault("quality_history", []).append({
        "date": _now, "brand": BRAND,
        "rate": _quality_rate, "collected": _detail_ok, "pending": _detail_pending,
    })
    # 质量历史也限 200 条
    if len(_prof["quality_history"]) > MAX_RUNS:
        _prof["quality_history"] = _prof["quality_history"][-MAX_RUNS:]

    # ── 配置记忆（下次运行时自动建议）──
    _prof["last_config"] = _run["config"]

    # ── 写入文件 ──
    with open(_PROFILE, "w", encoding="utf-8") as _f:
        json.dump(_prof, _f, ensure_ascii=False, indent=1)

    # ── 打印成长摘要 ──
    _top_brands = sorted(_prof["brands"].items(), key=lambda kv: kv[1]["runs"], reverse=True)[:3]
    _tb_str = ", ".join("%s(%d)" % (n, v["runs"]) for n, v in _top_brands)
    print("[grow] 第 %d 次运行 · 累计 %d 条 · 耗时 %.1fs · 常用: %s" %
          (_prof["total_runs"], _prof["total_items"], _elapsed, _tb_str or "无"))
    print("[grow] 数据质量: %d/%d 已采集(%.1f%%) · 待补采 %d" %
          (_detail_ok, _total_wids, _quality_rate, _detail_pending))
    if _b["runs"] > 1:
        print("[grow] %s 已跑 %d 次 · 上次 %s 条/%d 产品 · 平均耗时 %.1fs · 最佳质量 %.1f%%" %
              (BRAND, _b["runs"], _b["last_result"].get("items", 0),
               _b["last_result"].get("products", 0), _b["avg_elapsed"], _b["best_quality"]))

    # ── 价格趋势提示（从 sales_history 读取）──
    try:
        _hist_path = os.path.join(os.path.dirname(fn) if fn else ".", "sales_history.json")
        _hist = json.load(open(_hist_path, encoding="utf-8"))
        if len(_hist) >= 2:
            _prev = _hist[-2]
            _curr = _hist[-1]
            _prev_top1 = _prev["top10"][0] if _prev.get("top10") else {}
            _curr_top1 = _curr["top10"][0] if _curr.get("top10") else {}
            if _prev_top1 and _curr_top1 and _prev_top1.get("name") == _curr_top1.get("name"):
                _sales_delta = _curr_top1.get("total_sales", 0) - _prev_top1.get("total_sales", 0)
                _prev_prices = _prev_top1.get("lowest3_prices", [])
                _curr_prices = _curr_top1.get("lowest3_prices", [])
                if _prev_prices and _curr_prices:
                    _price_delta = round(_curr_prices[0] - _prev_prices[0], 2)
                    _sales_str = "+%d" % _sales_delta if _sales_delta >= 0 else str(_sales_delta)
                    _price_str = "+%.2f" % _price_delta if _price_delta >= 0 else str(_price_delta)
                    print("[trend] TOP1 销量变化: %s · 最低价变化: %s（对比 %s）" %
                          (_sales_str, _price_str, _prev.get("date", "")))
    except Exception:
        pass  # 趋势分析失败不影响出表

except Exception as _e:
    print("[grow] 成长记录写入失败（不影响出表）:", _e, file=sys.stderr)
